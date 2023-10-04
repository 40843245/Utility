import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

class RangeHandler():
    def __init__(self,rect: list[int] ):
        self.x = rect[0]
        self.y = rect[1]
        self.width = rect[2]
        self.height = rect[3]
        
class WorksheetHandler():
    def __init__(self):
        pass
    def SetWorksheet(self,worksheet):
        if worksheet == None:
            raise Exception("The specified worksheet is either none or empty. It is NOT allowed during change.")
        self.worksheet = worksheet
        self.rangeHandler = RangeHandler( [ 1,1,self.worksheet.max_column,self.worksheet.max_row] )
    def Search(self,searchText:str):
        r = WorksheetHandler.SearchInRange(self.worksheet,searchText,self.rangeHandler)
        return r
    @staticmethod
    def SearchInRange(worksheet : Worksheet ,searchText : str, rangeHandler:RangeHandler):
        wordInfo = list()
        for i in range(rangeHandler.y, rangeHandler.height + 1):
            for j in range(rangeHandler.x, rangeHandler.width + 1):
                val = worksheet.cell(i,j).value
                if searchText == val :      
                    wordInfo.append(worksheet.cell(i,j))
        return wordInfo
    
    def Replace(self,searchText : str, replaceText : str):
        r = WorksheetHandler.ReplaceInRange(self.worksheet,searchText,replaceText,self.rangeHandler)
        
    @staticmethod
    def ReplaceInRange(worksheet : Worksheet ,searchText : str, replaceText : str, rangeHandler:RangeHandler):
        wordInfo = list()
        for i in range(rangeHandler.y, rangeHandler.height + 1):
            for j in range(rangeHandler.x, rangeHandler.width + 1):
                val = worksheet.cell(i,j).value
                if searchText == val :      
                    wordInfo.append(worksheet.cell(i,j))
                    worksheet.cell(i,j).value = replaceText
        return wordInfo
        
class WorkbookHandler():
    def __init__(self):
        self.workbook = None
    def SetInputFullName(self,inputFullName: str):
        self.inputFullName = inputFullName
    def Load(self):
        if self.inputFullName == None:
            raise Exception("The specified full path name is either none or empty. It is NOT allowed.")
        self.workbook = openpyxl.load_workbook(self.inputFullName)
        if self.workbook == None:
            raise Exception("The workbook is either none or empty. It is NOT allowed.")
    def SetWorkbook(self , workbook):
        self.workbook = workbook
    def Search(self,searchText:str):
        resultInfo = list()
        worksheets = self.workbook.worksheets
        for i in range(0,len(worksheets),1):
            worksheetHandler = WorksheetHandler()
            worksheetHandler.SetWorksheet(worksheets[i])
            r = WorksheetHandler.Search(worksheetHandler,searchText)
            resultInfo.append(r)
        return resultInfo
    def Replace(self,searchText:str,replaceText):
        resultInfo = list()
        worksheets = self.workbook.worksheets
        for i in range(0,len(worksheets),1):
            worksheetHandler = WorksheetHandler()
            worksheetHandler.SetWorksheet(worksheets[i])
            r = WorksheetHandler.Replace(worksheetHandler,searchText,replaceText)
            resultInfo.append(r)
        self.AutoSave()
        return resultInfo
    def AutoSave(self):
        self.workbook.save(self.inputFullName)
class ExcelsHandler():
    def __init__(self):
        self.workbooks = None
    def SetInputFullNames(self,inputFullNames : list[str] ):
        self.inputFullNames = inputFullNames
    def Load(self):
        self.workbooksHandler = list()
        for i in range(0,len(self.inputFullNames),1):
            inputFullNames = self.inputFullNames[i]
            tempWorkbookHandler = WorkbookHandler()
            tempWorkbookHandler.SetInputFullName(inputFullNames)
            tempWorkbookHandler.Load()
            self.workbooksHandler.append(tempWorkbookHandler)
            del tempWorkbookHandler
    def Search(self,searchText:str):
        resultInfo = list()
        for i in range(0,len(self.workbooksHandler),1):
            r = WorkbookHandler.Search(self.workbooksHandler[i],searchText)
            resultInfo.append(r)
        return resultInfo
    def Replace(self,searchText:str,replaceText):
        resultInfo = list()
        for i in range(0,len(self.workbooksHandler),1):
            r = WorkbookHandler.Replace(self.workbooksHandler[i],searchText,replaceText)
            self.workbooksHandler[i].AutoSave()
            resultInfo.append(r)
        return resultInfo
    
if __name__ == '__main__':
    excelsHandler = ExcelsHandler()
    excelsHandler.SetInputFullNames([
            r"C:\Test\Demo\Demo1\inputs\Book2.xlsx","C:\Test\Demo\Demo1\inputs\Book3.xlsx"
         ])
    excelsHandler.Load()
    r = excelsHandler.Search("先輩がうざい後輩の話")
    print(r)
    r = excelsHandler.Replace("先輩がうざい後輩の話","前輩有夠煩")
    print(r)
